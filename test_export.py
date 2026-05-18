#!/usr/bin/env python
"""Test script to verify Excel export functionality"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Test data similar to SO2F2 Analyzer format
test_data = """Sr.No   Date     Time   Conc  Temp  Flow 
------- ---------- ------ ----- ----- ----- 
1       29/04/26   12:14  85.2  28.5  15.3
2       29/04/26   12:15  86.1  28.6  15.2
3       29/04/26   12:16  84.9  28.4  15.1"""

# Parse the data
lines = test_data.split('\n')
lines = [line for line in lines if line.strip()]

# Find data start
data_start_idx = 0
for idx, line in enumerate(lines):
    if line.strip().startswith('Sr.No') or (idx > 0 and '---' in lines[idx]):
        data_start_idx = idx + 2 if '---' in lines[idx] else idx + 1
        break

print(f"Data starts at index: {data_start_idx}")
print(f"Data lines: {lines[data_start_idx:]}")

# Detect delimiter
delimiter = None
for line in lines[data_start_idx:]:
    if '\t' in line:
        delimiter = '\t'
        break
    elif ',' in line:
        delimiter = ','
        break

print(f"Delimiter detected: {repr(delimiter)}")

# Parse data
parsed_data = []
max_cols = 0

for line in lines[data_start_idx:]:
    # Skip separator lines (lines with mostly dashes)
    if re.match(r'^\s*-+\s*(-\s*)*$', line.strip()):
        print(f"Skipping separator: {line.strip()}")
        continue
    
    if delimiter:
        cols = [col.strip() for col in line.split(delimiter) if col.strip()]
    else:
        # Split on multiple spaces
        cols = [col for col in re.split(r'\s{2,}|\s*\t\s*', line.strip()) if col.strip()]
    
    if cols:
        parsed_data.append(cols)
        max_cols = max(max_cols, len(cols))
        print(f"Parsed row: {cols}")

print(f"\nMax columns: {max_cols}")
print(f"Total data rows: {len(parsed_data)}")

# Create Excel file
try:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Data"
    
    # Add title
    ws['A1'] = "Test Export"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    # Add data
    row_num = 3
    for row_idx, row_data in enumerate(parsed_data):
        while len(row_data) < max_cols:
            row_data.append('')
        
        for col_idx, value in enumerate(row_data[:max_cols]):
            cell = ws.cell(row=row_num, column=col_idx + 1)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            if row_idx == 0:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            else:
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row_num += 1
    
    # Adjust column widths
    for col_idx in range(max_cols):
        ws.column_dimensions[chr(65 + col_idx)].width = 18
    
    # Save
    output_file = "d:\\spyder_pro\\p1\\test_export_output.xlsx"
    wb.save(output_file)
    print(f"\nSuccessfully created Excel file: {output_file}")
    
except Exception as e:
    print(f"\nError creating Excel file: {str(e)}")
    import traceback
    traceback.print_exc()
